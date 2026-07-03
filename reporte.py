"""Reporte mensual de contabilidad de videos por colaborador.

Lee el ledger `.videos_contabilidad.json` (lo escribe transcriber.py), arma un Excel
con el detalle y el resumen del MES ANTERIOR (o el mes que se le pase) y lo manda por
mail al diario. Pensado para dispararse el 1° de cada mes desde cron-job.org con
`main.py --videos-report`.
"""
import json
import smtplib
import ssl
import tempfile
from datetime import date, datetime
from email.message import EmailMessage
from email.utils import formataddr
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

from utils.config import get
from utils.logger import get_logger

logger = get_logger("reporte")

LEDGER = Path(__file__).parent / ".videos_contabilidad.json"

MESES = ["enero", "febrero", "marzo", "abril", "mayo", "junio",
         "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]


def _mes_anterior(hoy: date) -> str:
    y, m = (hoy.year - 1, 12) if hoy.month == 1 else (hoy.year, hoy.month - 1)
    return f"{y:04d}-{m:02d}"


def _leer_ledger() -> list[dict]:
    if not LEDGER.exists():
        return []
    try:
        return list(json.loads(LEDGER.read_text(encoding="utf-8-sig")))
    except Exception:
        return []


def _filas_del_mes(rows: list[dict], mes: str) -> list[dict]:
    """Filtra por mes (YYYY-MM) según fecha_recibido."""
    out = []
    for r in rows:
        fr = r.get("fecha_recibido", "")
        if fr[:7] == mes:
            out.append(r)
    return out


_PUBLICADOS = ("publicado", "publicado_solo_reel", "publicado_placa")


def _quien(r: dict) -> str:
    """Identidad del colaborador: el corresponsal de WhatsApp si lo hay; si no, el
    uploader de Drive."""
    return (r.get("corresponsal_nombre") or r.get("uploader") or "—").strip() or "—"


def _armar_excel(filas: list[dict], mes: str) -> Path:
    wb = openpyxl.Workbook()

    # Hoja Detalle
    det = wb.active
    det.title = "Detalle"
    cab = ["Fecha", "Colaborador", "Origen", "Volanta", "Título", "Estado", "Link"]
    det.append(cab)
    for c in det[1]:
        c.font = Font(bold=True)
    for r in sorted(filas, key=lambda x: x.get("fecha_recibido", "")):
        det.append([
            (r.get("fecha_recibido", "") or "")[:10],
            _quien(r),
            r.get("origen", "") or ("corresponsal-whatsapp" if r.get("corresponsal_nombre") else "drive"),
            r.get("volanta", ""),
            r.get("titulo", ""),
            r.get("estado", ""),
            r.get("post_url", ""),
        ])
    for col, ancho in zip("ABCDEFG", (12, 28, 20, 22, 50, 16, 45)):
        det.column_dimensions[col].width = ancho

    # Hoja Resumen (colaborador → cantidad)
    res = wb.create_sheet("Resumen")
    res.append(["Colaborador", "Videos enviados", "Publicados"])
    for c in res[1]:
        c.font = Font(bold=True)
    conteo: dict[str, list[int]] = {}
    for r in filas:
        u = _quien(r)
        conteo.setdefault(u, [0, 0])
        conteo[u][0] += 1
        if r.get("estado") in _PUBLICADOS:
            conteo[u][1] += 1
    for u, (total, pub) in sorted(conteo.items(), key=lambda kv: kv[1][0], reverse=True):
        res.append([u, total, pub])
    res.append([])
    res.append(["TOTAL", len(filas), sum(1 for r in filas if r.get("estado") in _PUBLICADOS)])
    for c in res[res.max_row]:
        c.font = Font(bold=True)
    for col, ancho in zip("ABC", (32, 16, 12)):
        res.column_dimensions[col].width = ancho

    # Hoja Colaboradores (la BASE DE DATOS del Programa de Corresponsales): solo los que
    # llegaron por WhatsApp, con su celular y conteo. Es lo que pide el spec del programa.
    corr = [r for r in filas if r.get("corresponsal_nombre")]
    if corr:
        col_sheet = wb.create_sheet("Colaboradores")
        col_sheet.append(["Nombre", "Celular", "Lugar", "Notas enviadas", "Publicadas", "Última fecha"])
        for c in col_sheet[1]:
            c.font = Font(bold=True)
        db: dict[str, dict] = {}
        for r in corr:
            nombre = (r.get("corresponsal_nombre") or "—").strip()
            d = db.setdefault(nombre, {"cel": "", "lugar": "", "tot": 0, "pub": 0, "fecha": ""})
            d["cel"] = r.get("corresponsal_celular") or d["cel"]
            d["lugar"] = r.get("corresponsal_lugar") or d["lugar"]
            d["tot"] += 1
            if r.get("estado") in _PUBLICADOS:
                d["pub"] += 1
            f = (r.get("fecha_recibido", "") or "")[:10]
            if f > d["fecha"]:
                d["fecha"] = f
        for nombre, d in sorted(db.items(), key=lambda kv: kv[1]["tot"], reverse=True):
            col_sheet.append([nombre, d["cel"], d["lugar"], d["tot"], d["pub"], d["fecha"]])
        for col, ancho in zip("ABCDEF", (28, 18, 24, 16, 12, 14)):
            col_sheet.column_dimensions[col].width = ancho

    salida = Path(tempfile.gettempdir()) / f"contabilidad_videos_{mes}.xlsx"
    wb.save(salida)
    logger.info(f"Excel armado: {salida} ({len(filas)} fila(s))")
    return salida


def _enviar(xlsx: Path, mes: str, cantidad: int) -> None:
    remitente = get("MAIL_FROM")
    password = get("MAIL_APP_PASSWORD")
    destino = get("VIDEOS_REPORT_EMAIL") or get("VIDEOS_NOTIFY_EMAIL") or remitente
    if not remitente or not password or not destino:
        logger.error("Faltan credenciales de mail (MAIL_FROM/MAIL_APP_PASSWORD): no se manda el reporte.")
        return
    host = get("SMTP_HOST") or "smtp.gmail.com"
    port = int(get("SMTP_PORT") or 587)
    nombre_from = get("MAIL_FROM_NAME") or "Diario La Campaña"

    y, m = mes.split("-")
    mes_largo = f"{MESES[int(m) - 1]} {y}"
    msg = EmailMessage()
    msg["From"] = formataddr((nombre_from, remitente))
    msg["To"] = destino
    msg["Subject"] = f"Contabilidad de videos — {mes_largo}"
    msg.set_content(
        f"Adjuntamos la contabilidad de videos de {mes_largo}.\n\n"
        f"Total de videos recibidos: {cantidad}.\n\n"
        f"En la hoja «Resumen» está el conteo por colaborador.\n\nDiario La Campaña"
    )
    msg.add_attachment(
        xlsx.read_bytes(),
        maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=xlsx.name,
    )
    try:
        ctx = ssl.create_default_context()
        with smtplib.SMTP(host, port, timeout=60) as server:
            server.starttls(context=ctx)
            server.login(remitente, password)
            server.send_message(msg)
        logger.info(f"Reporte enviado a {destino}")
    except Exception as e:
        logger.error(f"No se pudo enviar el reporte: {e}")


def run_videos_report(mes: str | None = None, dry_run: bool = False) -> None:
    mes = mes or _mes_anterior(date.today())
    logger.info(f"=== Reporte de videos del mes {mes} {'(dry-run)' if dry_run else ''} ===")

    rows = _leer_ledger()
    filas = _filas_del_mes(rows, mes)
    if not filas:
        logger.info(f"No hubo videos en {mes}. No se manda reporte.")
        return

    xlsx = _armar_excel(filas, mes)
    if dry_run:
        logger.info(f"[dry-run] Excel generado en {xlsx} ({len(filas)} filas). No se manda mail.")
        return
    _enviar(xlsx, mes, len(filas))
    logger.info("=== Reporte de videos: fin ===")

    # El 1° de cada mes, además del reporte, armamos el RANKING de corresponsales del mismo
    # mes (podio + premios). Va enganchado acá para reusar la corrida mensual que ya dispara
    # cron-job.org (--videos-report), sin crear otro cron. Best-effort: si falla, no rompe el reporte.
    try:
        from ranking import run_corresponsales_ranking
        run_corresponsales_ranking(mes=mes, dry_run=dry_run)
    except Exception as e:  # noqa: BLE001
        logger.error(f"El ranking de corresponsales falló (el reporte igual salió): {e}")
