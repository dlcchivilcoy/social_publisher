"""Lee los turnos de farmacias desde el correo del Colegio de Farmacéuticos
(farchivi@gmail.com) en la casilla dlc.chivilcoy@gmail.com, vía IMAP.

Dos tipos de correo:
  1) MENSUAL — asunto "TURNOS {MES} {AÑO}" con un Excel adjunto (TURNOS ...xlsx)
     que trae el cronograma del mes (día → 3 farmacias; las 2 primeras 24 hs,
     la última 8:30 a 22 hs).
  2) CAMBIO — asunto "TURNO CAMBIO HOY ..." con texto:
         LAS 24 HS.: ZURITA - PALUMBO
         HASTA LAS 22 HS.: ROSSI
     que reemplaza la terna de un día puntual.

Esto es más confiable que leer la imagen de dechivilcoy (OCR). Reusa las
credenciales de Gmail (MAIL_FROM / MAIL_APP_PASSWORD) que ya están en .env.
"""
import email
import imaplib
import io
import re
from datetime import date, datetime
from email.header import decode_header, make_header

from openpyxl import load_workbook

from utils.config import get
from utils.logger import get_logger

logger = get_logger("farmacias_mail")

MESES = ["ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
         "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]


def _remitente() -> str:
    return get("FARMACIAS_MAIL_FROM") or "farchivi@gmail.com"


def _dec(s: str) -> str:
    try:
        return str(make_header(decode_header(s or "")))
    except Exception:
        return s or ""


def _limpia(v) -> str:
    s = re.sub(r"\s+", " ", str(v or "").strip())
    return s.title()


def _txt(v) -> bool:
    return v is not None and str(v).strip() != ""


def _es_dia(v) -> bool:
    try:
        return 1 <= int(v) <= 31
    except (TypeError, ValueError):
        return False


# ── IMAP ───────────────────────────────────────────────────────────────────
def _conectar() -> imaplib.IMAP4_SSL:
    user = get("MAIL_FROM")
    pwd = get("MAIL_APP_PASSWORD")
    if not user or not pwd:
        raise RuntimeError("Faltan MAIL_FROM / MAIL_APP_PASSWORD en .env")
    M = imaplib.IMAP4_SSL("imap.gmail.com", 993, timeout=40)
    M.login(user, pwd)
    # "Todos" incluye archivados; si falla, INBOX.
    if M.select('"[Gmail]/Todos"', readonly=True)[0] != "OK":
        M.select("INBOX", readonly=True)
    return M


def _asuntos_recientes(M, n: int = 40) -> list[tuple[bytes, str]]:
    """Devuelve [(num, asunto)] de los últimos n correos del remitente (nuevo→viejo)."""
    typ, data = M.search(None, "FROM", _remitente())
    if typ != "OK" or not data or not data[0]:
        return []
    ids = data[0].split()[-n:]
    out = []
    for num in reversed(ids):
        typ, d = M.fetch(num, "(BODY.PEEK[HEADER.FIELDS (SUBJECT)])")
        subj = ""
        for item in d or []:
            if isinstance(item, tuple) and item[1]:
                subj = _dec(email.message_from_bytes(item[1]).get("Subject"))
        out.append((num, subj))
    return out


def _msg_completo(M, num):
    typ, d = M.fetch(num, "(RFC822)")
    return email.message_from_bytes(d[0][1])


def _texto_plano(msg) -> str:
    for part in msg.walk():
        disp = str(part.get("Content-Disposition") or "")
        if part.get_content_type() == "text/plain" and "attachment" not in disp:
            raw = part.get_payload(decode=True) or b""
            return raw.decode(part.get_content_charset() or "utf-8", "replace")
    return ""


def _xlsx_adjunto(msg) -> bytes | None:
    for part in msg.walk():
        fn = _dec(part.get_filename()) if part.get_filename() else ""
        if fn.lower().endswith(".xlsx"):
            return part.get_payload(decode=True)
    return None


# ── Parsers ──────────────────────────────────────────────────────────────────
def _parse_excel(xb: bytes):
    """Devuelve (anio, mes, {dia: [f1, f2, f3]}) a partir del Excel mensual."""
    wb = load_workbook(io.BytesIO(xb), data_only=True)
    ws = wb.active
    anio = mes = None
    dias: dict[int, list[str]] = {}
    for row in ws.iter_rows(values_only=True):
        cells = list(row) + [None] * max(0, 8 - len(row))
        if anio is None:
            for v in row:
                if isinstance(v, datetime):
                    anio, mes = v.year, v.month
                    break
        # Bloque izquierdo (días 1–15) y derecho (16–31).
        if _es_dia(cells[0]) and _txt(cells[1]):
            dias[int(cells[0])] = [_limpia(cells[1]), _limpia(cells[2]), _limpia(cells[3])]
        if _es_dia(cells[4]) and _txt(cells[5]):
            dias[int(cells[4])] = [_limpia(cells[5]), _limpia(cells[6]), _limpia(cells[7])]
    return anio, mes, dias


def _split_nombres(s: str) -> list[str]:
    s = (s or "").split("\n")[0]
    partes = re.split(r"[-,/]| y ", s)
    return [_limpia(p) for p in partes if _limpia(p)]


def _parse_cambio(texto: str, anio: int | None = None):
    """Devuelve (terna[list], fecha[date]) a partir del texto del cambio."""
    mfecha = re.search(r"(\d{1,2})\s*/\s*(\d{1,2})", texto)
    fecha = None
    if mfecha:
        dd, mm = int(mfecha.group(1)), int(mfecha.group(2))
        try:
            fecha = date(anio or date.today().year, mm, dd)
        except ValueError:
            fecha = None
    m24 = re.search(r"24\s*HS\.?\s*:\s*(.+)", texto, re.I)
    m22 = re.search(r"22\s*HS\.?\s*:\s*(.+)", texto, re.I)
    l24 = _split_nombres(m24.group(1)) if m24 else []
    l22 = _split_nombres(m22.group(1)) if m22 else []
    terna = l24 + l22  # las de 24 hs primero, la de 22 hs al final
    return (terna or None), fecha


# ── API pública ───────────────────────────────────────────────────────────────
def cronograma_mensual(anio: int, mes: int) -> dict | None:
    """{dia: [f1,f2,f3]} del mes pedido, leído del Excel del mail. None si no está."""
    M = _conectar()
    try:
        asuntos = _asuntos_recientes(M, 40)
        # 1) Por asunto "TURNOS {MES} {AÑO}".
        objetivo = f"TURNOS {MESES[mes - 1]} {anio}"
        for num, subj in asuntos:
            if subj and objetivo in subj.upper():
                xb = _xlsx_adjunto(_msg_completo(M, num))
                if xb:
                    a, m, dias = _parse_excel(xb)
                    if dias and m == mes:
                        logger.info(f"Cronograma de farmacias leído del mail: {subj} ({len(dias)} días)")
                        return dias
        # 2) Cualquier Excel cuyo mes/año coincida.
        for num, subj in asuntos:
            if subj and "TURNOS" in subj.upper():
                xb = _xlsx_adjunto(_msg_completo(M, num))
                if xb:
                    a, m, dias = _parse_excel(xb)
                    if dias and a == anio and m == mes:
                        logger.info(f"Cronograma de farmacias leído del mail (por fecha): {subj}")
                        return dias
    finally:
        try:
            M.logout()
        except Exception:
            pass
    return None


def cambio_del_dia(d: date) -> list[str] | None:
    """Terna de un cambio de turno para el día d, o None si no hay."""
    M = _conectar()
    try:
        for num, subj in _asuntos_recientes(M, 20):
            if subj and "CAMBIO" in subj.upper():
                terna, fecha = _parse_cambio(_texto_plano(_msg_completo(M, num)), d.year)
                if terna and fecha and (fecha.month, fecha.day) == (d.month, d.day):
                    logger.info(f"Cambio de turno del día detectado en el mail: {subj} → {terna}")
                    return terna
    finally:
        try:
            M.logout()
        except Exception:
            pass
    return None
